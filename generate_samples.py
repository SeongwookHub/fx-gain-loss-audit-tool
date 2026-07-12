# -*- coding: utf-8 -*-
"""
외환차손익/외화환산손익 검증 도구 설계용 샘플 데이터 생성기
- 분개장.xlsx, 명세서(외화자산부채명세서).xlsx, 계정별원장.xlsx 3종 생성
- 8개 거래(TXN001~TXN008)에 실현/미실현, 정상/오류 케이스를 섞어 놓음
- 결산일: 2025-12-31 (12월 결산법인 가정)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BOLD = Font(name=FONT_NAME, bold=True)
NORMAL = Font(name=FONT_NAME)

# ------------------------------------------------------------------
# 거래 마스터 데이터 (8건) - 여기서 한 번만 정의하고 3개 파일에서 재사용
# ------------------------------------------------------------------
# 각 거래: 거래ID, 거래처, 통화, 계정성격(채권/채무), 계정과목,
#          발생일, 발생외화금액, 발생적용환율,
#          결제일(None=기말 미결제), 결제적용환율(None 가능),
#          기말재평가여부, 기말재평가적용환율(None 가능),
#          비고(참고용 - 의도된 오류 설명, 실제 회사 파일에는 없는 내용이므로
#              분개장/명세서에는 넣지 않고 별도 참고 시트에만 기록)

TXNS = [
    dict(id="TXN001", partner="ABC Trading", currency="USD", side="채권",
         account="108 외상매출금(외화)", counter_account="401 매출",
         occur_date="2025-03-10", occur_fc=10000, occur_rate=1446.30,   # 실제 매매기준율
         settle_date="2025-04-25", settle_rate=1430.20,                # 실제 매매기준율
         yearend=False, yearend_rate=None,
         note="정상 케이스 - 실제 환율 그대로 정확히 계상됨(외환차손 161,000원, USD 하락)"),

    dict(id="TXN002", partner="Global Parts Co", currency="USD", side="채권",
         account="108 외상매출금(외화)", counter_account="401 매출",
         occur_date="2025-02-05", occur_fc=20000, occur_rate=1460.80,  # 실제 매매기준율
         settle_date="2025-05-15", settle_rate=1438.50,  # 오류: 실제 결제일(1415.80) 대신 전월말(2025-04-30) 환율을 잘못 적용
         yearend=False, yearend_rate=None,
         note="오류 케이스 - 결제일 실제 환율(1415.80)이 아닌 전월말 환율(1438.50)을 잘못 적용. "
              "실제 데이터 기준 괴리율은 1.6%로 5% 개별기준은 통과하지만, 합계중요성에서 걸림"),

    dict(id="TXN003", partner="Osaka Parts KK", currency="JPY", side="채권",
         account="108 외상매출금(외화)", counter_account="401 매출",
         occur_date="2025-05-30", occur_fc=1000000, occur_rate=959.27,  # 실제 매매기준율(100엔당)
         settle_date="2025-07-18", settle_rate=937.73,  # 실제 매매기준율(100엔당) - 정상이라면 /100 해서 적용
         settle_raw_no_divide=True,  # 오류: 100 나누기를 누락하고 937.73을 그대로 곱함
         yearend=False, yearend_rate=None,
         note="오류 케이스 - JPY는 100엔 단위 고시인데 100 나누기를 누락해 약 100배 과대계상"),

    dict(id="TXN004", partner="Hanaro Export", currency="USD", side="채권",
         account="108 외상매출금(외화)", counter_account="401 매출",
         occur_date="2025-08-08", occur_fc=5000, occur_rate=1384.20,   # 실제 매매기준율
         settle_date="2025-09-30", settle_rate=1402.20,                # 실제 매매기준율
         yearend=False, yearend_rate=None,
         note="오류 케이스 - 환율 자체는 실제값을 정확히 썼지만, 외환차익 90,000원을 별도 계정으로 "
              "인식하지 않고 채권 계정에서 바로 상계(외환차익 계상 누락). 괴리율 0%라 1단계에선 안 걸리고 "
              "합계중요성에서 걸리는 케이스",
         omit_gain_loss_line=True),

    dict(id="TXN005", partner="Berlin Materials GmbH", currency="EUR", side="채무",
         account="251 외상매입금(외화)", counter_account="146 원재료매입",
         occur_date="2025-04-02", occur_fc=3000, occur_rate=1589.14,   # 실제 매매기준율
         settle_date="2025-06-11", settle_rate=1554.64,                # 실제 매매기준율
         yearend=False, yearend_rate=None,
         note="정상 케이스 - 실제 환율 그대로 정확히 계상됨(외환차익 103,500원, EUR 하락으로 부채 상환액 감소)"),

    dict(id="TXN006", partner="Pacific Foods Inc", currency="USD", side="채권",
         account="108 외상매출금(외화)", counter_account="401 매출",
         occur_date="2025-11-05", occur_fc=8000, occur_rate=1437.70,   # 실제 매매기준율
         settle_date=None, settle_rate=None,
         yearend=True, yearend_rate=1434.90,  # 실제 2025-12-31 매매기준율
         note="정상 케이스 - 기말 미결제, 실제 결산일 환율로 정확히 재평가(외화환산손실 22,400원)"),

    dict(id="TXN007", partner="Nagoya Steel Co", currency="JPY", side="채무",
         account="251 외상매입금(외화)", counter_account="146 원재료매입",
         occur_date="2025-10-20", occur_fc=2000000, occur_rate=940.66,  # 실제 매매기준율(100엔당)
         settle_date=None, settle_rate=None,
         yearend=False, yearend_rate=None,  # 재평가 자체를 누락
         note="오류 케이스 - 기말 미결제 상태인데 외화환산손익 재평가 분개 자체가 누락됨(완전성 이슈)"),

    dict(id="TXN008", partner="Texas Equip LLC", currency="USD", side="채권",
         account="108 외상매출금(외화)", counter_account="401 매출",
         occur_date="2025-09-05", occur_fc=15000, occur_rate=1392.70,  # 실제 매매기준율
         settle_date=None, settle_rate=None,
         yearend=True, yearend_rate=1470.00,  # 오류: 실제 당기말(1434.90) 대신 전기말(2024-12-31) 환율을 잘못 사용
         note="오류 케이스 - 기말 재평가 시 당기말 환율(1434.90)이 아닌 전기말 환율(1470.00)을 잘못 사용해 "
              "외화환산이익을 과대계상(1,159,500원 vs 정상 633,000원)"),
]

FX_ACCOUNT_MAP = {
    "USD": "USD", "EUR": "EUR", "JPY": "JPY(100)",
}


def krw(fc, rate, is_jpy=False):
    """외화금액 x 환율 = 원화금액. JPY는 100엔당 고시이므로 100으로 나눔."""
    r = rate / 100 if is_jpy else rate
    return round(fc * r)


# ------------------------------------------------------------------
# 1. 분개장.xlsx
# ------------------------------------------------------------------

def build_journal():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "분개장"

    headers = ["전표번호(거래ID)", "라인번호", "일자", "구분", "계정코드", "계정과목",
               "차변금액", "대변금액", "통화", "외화금액", "적용환율", "거래처", "적요"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    row = 2
    for t in TXNS:
        is_jpy = t["currency"] == "JPY"
        acct_code, acct_name = t["account"].split(" ", 1)
        counter_code, counter_name = t["counter_account"].split(" ", 1)

        # --- 발생 전표 (2줄: 채권/채무 계정 vs 매출/매입 상대계정) ---
        occur_krw = krw(t["occur_fc"], t["occur_rate"], is_jpy)
        if t["side"] == "채권":
            ws.append([t["id"], 1, t["occur_date"], "발생", acct_code, acct_name,
                       occur_krw, "", t["currency"], t["occur_fc"], t["occur_rate"],
                       t["partner"], f"{t['partner']} 외상매출 발생"])
            ws.append([t["id"], 2, t["occur_date"], "발생", counter_code, counter_name,
                       "", occur_krw, t["currency"], t["occur_fc"], t["occur_rate"],
                       t["partner"], f"{t['partner']} 매출 인식"])
        else:  # 채무
            ws.append([t["id"], 1, t["occur_date"], "발생", counter_code, counter_name,
                       occur_krw, "", t["currency"], t["occur_fc"], t["occur_rate"],
                       t["partner"], f"{t['partner']} 원재료매입 발생"])
            ws.append([t["id"], 2, t["occur_date"], "발생", acct_code, acct_name,
                       "", occur_krw, t["currency"], t["occur_fc"], t["occur_rate"],
                       t["partner"], f"{t['partner']} 외상매입금 발생"])
        row += 2

        # --- 결제 전표 (실현손익 있는 경우) ---
        if t["settle_date"]:
            if t.get("settle_raw_no_divide"):
                # 오류 재현: 100엔당 고시값을 그대로 1엔당인 것처럼 곱함 (100 나누기 누락)
                settle_krw = round(t["occur_fc"] * t["settle_rate"])
            else:
                settle_krw = krw(t["occur_fc"], t["settle_rate"], is_jpy)
            diff = settle_krw - occur_krw  # 채권: 결제-발생 / 채무: 부호 반대 적용
            if t["side"] == "채권":
                # 외화예금 차변, 채권 대변(장부가), 차익/차손
                ws.append([t["id"], 3, t["settle_date"], "결제", "103", "외화예금",
                           settle_krw, "", t["currency"], t["occur_fc"], t["settle_rate"],
                           t["partner"], f"{t['partner']} 대금 회수"])
                if t.get("omit_gain_loss_line"):
                    # 오류: 외환차익 계정을 쓰지 않고 채권 계정에서 결제환율 기준으로 그대로 상계
                    ws.append([t["id"], 4, t["settle_date"], "결제", acct_code, acct_name,
                               "", settle_krw, t["currency"], t["occur_fc"], t["settle_rate"],
                               t["partner"], f"{t['partner']} 채권 상계(외환차익 별도 인식 없음)"])
                else:
                    ws.append([t["id"], 4, t["settle_date"], "결제", acct_code, acct_name,
                               "", occur_krw, t["currency"], t["occur_fc"], t["occur_rate"],
                               t["partner"], f"{t['partner']} 채권 상계(장부가)"])
                    if diff != 0:
                        gain_acct = ("907", "외환차익") if diff > 0 else ("957", "외환차손")
                        if diff > 0:
                            ws.append([t["id"], 5, t["settle_date"], "결제", gain_acct[0], gain_acct[1],
                                       "", abs(diff), t["currency"], "", "", t["partner"],
                                       f"{t['partner']} 외환차익 인식"])
                        else:
                            ws.append([t["id"], 5, t["settle_date"], "결제", gain_acct[0], gain_acct[1],
                                       abs(diff), "", t["currency"], "", "", t["partner"],
                                       f"{t['partner']} 외환차손 인식"])
            else:  # 채무 결제
                ws.append([t["id"], 3, t["settle_date"], "결제", acct_code, acct_name,
                           occur_krw, "", t["currency"], t["occur_fc"], t["occur_rate"],
                           t["partner"], f"{t['partner']} 채무 상계(장부가)"])
                ws.append([t["id"], 4, t["settle_date"], "결제", "103", "외화예금",
                           "", settle_krw, t["currency"], t["occur_fc"], t["settle_rate"],
                           t["partner"], f"{t['partner']} 대금 지급"])
                diff_liab = settle_krw - occur_krw  # 부채 상환시 더 많이 나가면 손실
                if diff_liab != 0:
                    if diff_liab > 0:
                        ws.append([t["id"], 5, t["settle_date"], "결제", "957", "외환차손",
                                   abs(diff_liab), "", t["currency"], "", "", t["partner"],
                                   f"{t['partner']} 외환차손 인식"])
                    else:
                        ws.append([t["id"], 5, t["settle_date"], "결제", "907", "외환차익",
                                   "", abs(diff_liab), t["currency"], "", "", t["partner"],
                                   f"{t['partner']} 외환차익 인식"])

        # --- 기말 재평가 전표 (yearend=True인 경우만; 미실현 손익) ---
        if t["yearend"]:
            ye_krw = krw(t["occur_fc"], t["yearend_rate"], is_jpy)
            ye_diff = ye_krw - occur_krw
            if t["side"] == "채권":
                if ye_diff != 0:
                    if ye_diff > 0:
                        ws.append([t["id"], 3, "2025-12-31", "기말평가", acct_code, acct_name,
                                   abs(ye_diff), "", t["currency"], t["occur_fc"], t["yearend_rate"],
                                   t["partner"], f"{t['partner']} 외화환산이익 인식"])
                        ws.append([t["id"], 4, "2025-12-31", "기말평가", "908", "외화환산이익",
                                   "", abs(ye_diff), t["currency"], "", "", t["partner"],
                                   f"{t['partner']} 외화환산이익 인식"])
                    else:
                        ws.append([t["id"], 3, "2025-12-31", "기말평가", "958", "외화환산손실",
                                   abs(ye_diff), "", t["currency"], "", "", t["partner"],
                                   f"{t['partner']} 외화환산손실 인식"])
                        ws.append([t["id"], 4, "2025-12-31", "기말평가", acct_code, acct_name,
                                   "", abs(ye_diff), t["currency"], t["occur_fc"], t["yearend_rate"],
                                   t["partner"], f"{t['partner']} 외화환산손실 인식"])
            # 채무 기말평가는 이번 샘플엔 없음(TXN007이 재평가 누락 케이스)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"

    return wb


# ------------------------------------------------------------------
# 2. 명세서(외화자산부채명세서).xlsx
# ------------------------------------------------------------------

def build_schedule():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "외화자산부채명세서"

    headers = ["전표번호(거래ID)", "거래처", "통화", "구분(채권/채무)", "계정과목",
               "발생일", "발생외화금액", "발생환율", "발생원화금액",
               "결제일", "결제외화금액", "결제환율", "결제원화금액",
               "기말미결제외화잔액", "비고"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    for t in TXNS:
        is_jpy = t["currency"] == "JPY"
        occur_krw = krw(t["occur_fc"], t["occur_rate"], is_jpy)
        if t["settle_date"]:
            if t.get("settle_raw_no_divide"):
                settle_krw = round(t["occur_fc"] * t["settle_rate"])
            else:
                settle_krw = krw(t["occur_fc"], t["settle_rate"], is_jpy)
        else:
            settle_krw = ""
        outstanding_fc = t["occur_fc"] if not t["settle_date"] else 0
        note = "기말 현재 미결제" if not t["settle_date"] else "기중 결제 완료"
        ws.append([
            t["id"], t["partner"], t["currency"], t["side"], t["account"],
            t["occur_date"], t["occur_fc"], t["occur_rate"], occur_krw,
            t["settle_date"] or "", t["occur_fc"] if t["settle_date"] else "",
            t["settle_rate"] or "", settle_krw,
            outstanding_fc, note,
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"
    return wb


# ------------------------------------------------------------------
# 3. 계정별원장.xlsx
# ------------------------------------------------------------------

def build_ledger():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "계정별원장"

    headers = ["계정코드", "계정과목", "기초잔액", "차변합계", "대변합계", "기말잔액"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    # 분개장 로직을 그대로 재현하여 계정별 차/대변 합계 산출
    debit = {}
    credit = {}

    def add(acc, code_name, d, c):
        code, name = code_name.split(" ", 1)
        key = (code, name)
        debit[key] = debit.get(key, 0) + d
        credit[key] = credit.get(key, 0) + c

    for t in TXNS:
        is_jpy = t["currency"] == "JPY"
        acct = t["account"]
        counter = t["counter_account"]
        occur_krw = krw(t["occur_fc"], t["occur_rate"], is_jpy)

        if t["side"] == "채권":
            add(acct, acct, occur_krw, 0)
            add(counter, counter, 0, occur_krw)
        else:
            add(counter, counter, occur_krw, 0)
            add(acct, acct, 0, occur_krw)

        if t["settle_date"]:
            if t.get("settle_raw_no_divide"):
                settle_krw = round(t["occur_fc"] * t["settle_rate"])
            else:
                settle_krw = krw(t["occur_fc"], t["settle_rate"], is_jpy)
            diff = settle_krw - occur_krw
            if t["side"] == "채권":
                add("103 외화예금", "103 외화예금", settle_krw, 0)
                if t.get("omit_gain_loss_line"):
                    add(acct, acct, 0, settle_krw)
                else:
                    add(acct, acct, 0, occur_krw)
                    if diff > 0:
                        add("907 외환차익", "907 외환차익", 0, diff)
                    elif diff < 0:
                        add("957 외환차손", "957 외환차손", abs(diff), 0)
            else:
                add(acct, acct, occur_krw, 0)
                add("103 외화예금", "103 외화예금", 0, settle_krw)
                diff_liab = settle_krw - occur_krw
                if diff_liab > 0:
                    add("957 외환차손", "957 외환차손", abs(diff_liab), 0)
                elif diff_liab < 0:
                    add("907 외환차익", "907 외환차익", 0, abs(diff_liab))

        if t["yearend"]:
            ye_krw = krw(t["occur_fc"], t["yearend_rate"], is_jpy)
            ye_diff = ye_krw - occur_krw
            if t["side"] == "채권" and ye_diff != 0:
                if ye_diff > 0:
                    add(acct, acct, abs(ye_diff), 0)
                    add("908 외화환산이익", "908 외화환산이익", 0, abs(ye_diff))
                else:
                    add("958 외화환산손실", "958 외화환산손실", abs(ye_diff), 0)
                    add(acct, acct, 0, abs(ye_diff))

    # 기초잔액 (전기이월 - 8건 거래와 무관한 별도 잔액, 실무 반영을 위해 가정)
    opening = {
        ("103", "외화예금"): 50000000,
        ("108", "외상매출금(외화)"): 12000000,
        ("251", "외상매입금(외화)"): 8000000,
    }

    # 계정별원장에 '완전성 이슈'를 하나 심어둠: 외환차익 계정 기말잔액을
    # 분개장 합계보다 50,000원 더 크게 표시 (상위 시스템 수기조정분이 분개장 상세에는
    # 반영되지 않은 것으로 가정 - 3개 문서 교차검증이 필요한 이유)
    LEDGER_ADJUSTMENT = {("907", "외환차익"): 50000}

    all_keys = set(list(debit.keys()) + list(credit.keys()) + list(opening.keys()))
    # 계정 정렬 순서 고정
    order = ["103", "108", "251", "401", "146", "907", "957", "908", "958"]
    all_keys = sorted(all_keys, key=lambda k: order.index(k[0]) if k[0] in order else 99)

    for key in all_keys:
        code, name = key
        op = opening.get(key, 0)
        d = debit.get(key, 0)
        c = credit.get(key, 0)
        adj = LEDGER_ADJUSTMENT.get(key, 0)
        c += adj  # 조정분 반영(대변 성격 계정 가정)
        # 자산/비용성 계정은 차변 증가, 부채/수익성 계정은 대변 증가 기준으로 기말잔액 계산
        if code in ("103", "108", "146", "957", "958"):
            ending = op + d - c
        else:
            ending = op + c - d
        ws.append([code, name, op, d, c, ending])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"
    return wb


# ------------------------------------------------------------------
# 4. 참고용 시트(검증 포인트 정리) - 실제 회사 파일에는 없는, 성욱님 검토용
# ------------------------------------------------------------------

def build_answer_key():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검증포인트(참고용)"
    ws.append(["거래ID", "유형", "정상/오류", "설명"])
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    for t in TXNS:
        유형 = "기말 미실현(외화환산)" if t["yearend"] or (not t["settle_date"] and not t["yearend"]) else "실현(결제완료)"
        정상오류 = "오류" if "오류" in t["note"] else "정상"
        ws.append([t["id"], 유형, 정상오류, t["note"]])

    ws.append([])
    ws.append(["계정별원장 관련 별도 이슈"])
    ws.append(["907 외환차익 계정", "-", "오류",
               "계정별원장 기말잔액이 분개장 합계보다 50,000원 많음(상위 시스템 수기조정분 미반영 가정) - 3개 문서 교차검증 필요성을 보여주는 케이스"])

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 30
    ws.column_dimensions["D"].width = 70
    return wb


if __name__ == "__main__":
    build_journal().save("분개장.xlsx")
    build_schedule().save("명세서_외화자산부채명세서.xlsx")
    build_ledger().save("계정별원장.xlsx")
    build_answer_key().save("검증포인트_참고용.xlsx")
    print("생성 완료")
