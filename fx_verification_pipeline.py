# -*- coding: utf-8 -*-
"""
외환차손익 / 외화환산손익 검증 파이프라인 (1단계 스크리닝 + 2단계 증빙 OCR)
================================================================

[전체 구조]

  A. 외환차손익(결제 건, 분개장의 "결제" 라인)
     1단계: 전수 스크리닝 - 회사 적용환율 vs 그날 공식 매매기준율(수출입은행 API)
            괴리율이 5%를 넘는 건만 자동 추출 (전수조사가 불가능한 현실을 반영,
            "어떤 걸 표본으로 볼지"를 사람이 감으로 고르지 않고 수치 기준으로 자동 추출)
     2단계: 1단계에서 걸린 건만 증빙(은행 외환거래확인서/SWIFT/외화예금 거래명세표 등
            스캔 이미지)을 Claude Vision으로 읽어 실제 적용환율을 추출하고,
            그 환율로 재계산한 외환차손익을 회사 계상액과 최종 비교

  B. 외화환산손익(기말평가 건, 분개장의 "기말평가" 라인)
     결산일 하나의 공식 환율로 전수 재계산 (거래별로 증빙이 다를 이유가 없으므로
     OCR 2단계가 필요 없음 - 전수 자동 검증)

[사용 전 준비]
  1. data.go.kr에서 한국수출입은행 환율 API 인증키 발급 (EXIM_AUTH_KEY 환경변수)
  2. 증빙 이미지가 있다면 evidence/ 폴더에 "{거래ID}.png" 또는 "{거래ID}.jpg" 형태로 저장
     (예: evidence/TXN002.png) - 없으면 2단계는 "증빙 요청 필요"로 표시만 하고 넘어감
  3. Claude API 사용을 위해 ANTHROPIC_API_KEY 환경변수 설정 (2단계 OCR용)
  4. pip install anthropic pandas openpyxl requests --break-system-packages
"""

import os
import re
import json
import time
import base64
from datetime import datetime, timedelta

import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# 0. 설정
# ------------------------------------------------------------------

EXIM_AUTH_KEY = os.environ.get("EXIM_AUTH_KEY", "여기에_발급받은_인증키_입력")
EXIM_BASE_URL = "https://oapi.koreaexim.go.kr/site/program/financial/exchangeJSON"
# 참고: 2025.6.25부로 요청 URL 도메인이 www.koreaexim.go.kr -> oapi.koreaexim.go.kr로 변경됨.
# 기존 도메인(www.koreaexim.go.kr)은 점진적으로 종료 예정이라 응답이 없거나 타임아웃 날 수 있음.

TOLERANCE_PCT = 0.05          # 1단계 이상치 판단 기준: 괴리율 5%
GAIN_LOSS_TOLERANCE_KRW = 1000  # 재계산 금액과 회사 계상액의 허용 오차(원 단위 반올림 차이)
AMPT = float(os.environ.get("FX_AMPT", 3000000))  # 허용가능 오류금액(Tolerable Misstatement) - 감사팀 산정치로 교체

CUR_UNIT_MAP = {"USD": "USD", "JPY": "JPY(100)", "EUR": "EUR", "CNH": "CNH"}

EVIDENCE_DIR = "evidence"      # 증빙 이미지 폴더

_rate_cache = {}


# ------------------------------------------------------------------
# 1. 수출입은행 환율 API
# ------------------------------------------------------------------

def fetch_rates_for_date(date_str: str, _retries: int = 2) -> dict:
    """특정 날짜(YYYYMMDD)의 전체 통화 매매기준율표를 API에서 가져옴. RESULT 코드 체크 포함.
    일시적 타임아웃/연결 오류는 짧게 대기 후 재시도(최대 2회)."""
    if date_str in _rate_cache:
        return _rate_cache[date_str]

    params = {"authkey": EXIM_AUTH_KEY, "data": "AP01", "searchdate": date_str}

    for attempt in range(_retries + 1):
        try:
            resp = requests.get(EXIM_BASE_URL, params=params, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            break
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            if attempt < _retries:
                time.sleep(1.5 * (attempt + 1))  # 1.5초, 3초 순으로 대기 후 재시도
                continue
            raise

    if not data:
        # 주말/공휴일 등으로 데이터가 없는 경우 (빈 리스트) - 폴백 대상
        _rate_cache[date_str] = {}
        return {}

    rate_table = {}
    for row in data:
        result = str(row.get("result", ""))
        if result == "2":
            raise ValueError("수출입은행 API 오류(RESULT=2): DATA 코드 오류")
        elif result == "3":
            raise ValueError("수출입은행 API 오류(RESULT=3): 인증코드(authkey) 오류 - 키를 확인하세요")
        elif result == "4":
            raise ValueError("수출입은행 API 오류(RESULT=4): 일일 호출 제한 초과 - 내일 다시 시도하거나 캐시를 활용하세요")
        try:
            rate_table[row["cur_unit"]] = float(row["deal_bas_r"].replace(",", ""))
        except (KeyError, ValueError):
            continue

    _rate_cache[date_str] = rate_table
    time.sleep(0.15)  # 짧은 pacing 지연 - 짧은 시간에 요청이 몰려 서버가 지연/거부하는 것을 예방
    return rate_table


def get_official_rate(date_obj: datetime, cur_unit: str, max_lookback_days: int = 5) -> tuple:
    """해당일 공식 매매기준율. 주말/공휴일이면 직전 영업일로 폴백.
    반환값: (환율, 실제 적용된 날짜) - 폴백된 경우 실제 날짜를 알아야 리포트에 남길 수 있음"""
    for i in range(max_lookback_days + 1):
        d = date_obj - timedelta(days=i)
        rate_table = fetch_rates_for_date(d.strftime("%Y%m%d"))
        if cur_unit in rate_table:
            return rate_table[cur_unit], d
    raise ValueError(f"{date_obj.date()} 기준 {cur_unit} 환율을 찾을 수 없습니다 (휴장일 {max_lookback_days}일 초과)")


def normalize_rate(raw_rate: float, currency: str) -> float:
    """JPY는 100엔당 고시이므로 1엔당 단가로 환산."""
    return raw_rate / 100 if currency == "JPY" else raw_rate


# ------------------------------------------------------------------
# 2. 분개장 파싱 - 결제 건 / 기말평가 건 추출
# ------------------------------------------------------------------

REQUIRED_JOURNAL_COLUMNS = ["전표번호(거래ID)", "일자", "구분", "계정코드", "계정과목"]


def _detect_header_row(path: str, max_scan_rows: int = 10) -> int:
    """실제 회사 파일에는 제목행/빈행이 표 위에 붙어있는 경우가 흔하다.
    필수 컬럼명이 실제로 나타나는 행을 찾아서 그 행을 헤더로 사용한다.
    못 찾으면 0(첫 행)을 그대로 반환해 기존 동작을 유지한다."""
    preview = pd.read_excel(path, header=None, nrows=max_scan_rows)
    for i in range(len(preview)):
        row_values = set(str(v).strip() for v in preview.iloc[i].tolist())
        if all(col in row_values for col in REQUIRED_JOURNAL_COLUMNS):
            return i
    return 0


def load_journal(path: str) -> pd.DataFrame:
    header_row = _detect_header_row(path)
    df = pd.read_excel(path, header=header_row)
    df = df.dropna(how="all")  # 완전히 빈 행 제거 (제목/구분용 빈 행 대응)

    # 계정코드가 "108.0", " 108" 등으로 지저분하게 들어와도 비교 가능하도록 정규화
    df["계정코드"] = df["계정코드"].apply(_normalize_account_code)
    # 계정과목명 앞뒤 공백 제거 (ERP export에서 흔함)
    df["계정과목"] = df["계정과목"].astype(str).str.strip()
    # 통화 코드 대소문자 통일
    df["통화"] = df["통화"].astype(str).str.strip().str.upper()
    # 외화금액이 "15,000" 같은 콤마 포함 텍스트로 들어온 경우 숫자로 변환
    if "외화금액" in df.columns:
        df["외화금액"] = (
            df["외화금액"].astype(str).str.replace(",", "", regex=False).replace("nan", None)
        )
        df["외화금액"] = pd.to_numeric(df["외화금액"], errors="coerce")

    df["일자"] = pd.to_datetime(df["일자"])
    return df


def _normalize_account_code(code) -> str:
    """108, "108", "108.0", " 108 " 등 다양한 형태로 들어오는 계정코드를
    일관된 문자열 형태("108")로 정규화."""
    try:
        return str(int(float(str(code).strip())))
    except (ValueError, TypeError):
        return str(code).strip()


def extract_settlement_transactions(journal_df: pd.DataFrame) -> pd.DataFrame:
    """거래ID+결제일 단위로 '결제' 구분 라인을 묶어 결제 이벤트를 요약.
    한 거래가 여러 날짜에 나눠 결제된 경우(분할결제)에도 각 결제일을 별도
    이벤트로 인식한다 (거래ID만으로 묶으면 분할결제 중 일부가 누락되거나
    손익이 뒤섞이는 문제가 있었음)."""
    settle_rows = journal_df[(journal_df["구분"] == "결제") & (journal_df["외화금액"].notna())].copy()

    records = []
    for (txn_id, settle_date), group in settle_rows.groupby(["전표번호(거래ID)", "일자"]):
        # 채권/채무 결제 모두 '외화예금' 라인이 결제 시점 실제 자금 흐름의 기준
        cash_row = group[group["계정과목"] == "외화예금"]
        row = cash_row.iloc[0] if not cash_row.empty else group.iloc[0]
        currency = row["통화"]
        fc_amount = row["외화금액"]
        partner = row["거래처"]

        # 적용환율 '컬럼값'을 그대로 믿지 않고, 실제 분개된 원화금액에서 환율을 역산.
        booked_krw = row[["차변금액", "대변금액"]].fillna(0).abs().max()
        implied_rate = booked_krw / fc_amount if fc_amount else None

        # 907(외환차익)/957(외환차손) 계정 금액은 '같은 결제일 그룹' 안에서만 찾는다
        # (분할결제의 경우 결제 건마다 손익 라인이 따로 있으므로, 거래ID 전체를 보면
        #  서로 다른 결제 건의 손익이 뒤섞인다)
        # 907(외환차익)/957(외환차손) 라인은 외화금액이 비어있는(NaN) 경우가 많아
        # 앞서 외화금액 기준으로 필터링된 group(=settle_rows)에는 안 잡힐 수 있다.
        # journal_df 원본에서 같은 거래ID + 같은 결제일로 다시 조회해야 누락이 없다.
        same_leg = journal_df[(journal_df["전표번호(거래ID)"] == txn_id) &
                               (journal_df["구분"] == "결제") &
                               (journal_df["일자"] == settle_date)]
        gain_row = same_leg[same_leg["계정과목"] == "외환차익"]
        loss_row = same_leg[same_leg["계정과목"] == "외환차손"]

        booked_gain_loss = 0
        if not gain_row.empty:
            booked_gain_loss = gain_row.iloc[0][["차변금액", "대변금액"]].fillna(0).abs().max()
        elif not loss_row.empty:
            booked_gain_loss = -loss_row.iloc[0][["차변금액", "대변금액"]].fillna(0).abs().max()

        # 채권/채무 구분 - 거래ID 전체의 발생 라인 기준 (분할결제라도 발생은 1건)
        occur_rows = journal_df[(journal_df["전표번호(거래ID)"] == txn_id) &
                                 (journal_df["구분"] == "발생") &
                                 (journal_df["계정코드"].isin(["108", "251"]))]
        side = "채권" if not occur_rows.empty and occur_rows.iloc[0]["계정코드"] == "108" else "채무"

        # 데이터 무결성 경고: 같은 거래ID인데 통화나 거래처가 서로 다른 '발생' 라인이
        # 여러 개 섞여 있으면, 전표번호가 중복 채번되어 서로 다른 거래가 합쳐졌을 가능성이
        # 높다. 이 경우 아래 occur_krw는 첫 번째 라인 기준으로만 계산되므로 신뢰할 수 없다.
        integrity_warning = None
        if occur_rows["통화"].nunique() > 1 or occur_rows["거래처"].nunique() > 1:
            integrity_warning = (
                f"전표번호 중복 의심 - 동일 거래ID({txn_id})에 서로 다른 통화/거래처의 "
                f"발생 라인이 {len(occur_rows)}건 섞여 있음. 아래 금액은 신뢰할 수 없으니 원본 확인 필요"
            )

        # 발생원화금액(occur_krw)은 '이 결제 건'에서 실제로 상계된 채권/채무 장부금액을 사용.
        clear_row = group[group["계정코드"].isin(["108", "251"])]
        if not clear_row.empty:
            occur_krw = clear_row.iloc[0][["차변금액", "대변금액"]].fillna(0).abs().max()
        else:
            occur_krw = occur_rows.iloc[0][["차변금액", "대변금액"]].fillna(0).abs().max() if not occur_rows.empty else None

        records.append({
            "거래ID": txn_id, "결제일": settle_date, "통화": currency, "구분": side, "거래처": partner,
            "데이터무결성경고": integrity_warning,
            "외화금액": fc_amount, "결제원화금액": booked_krw, "발생원화금액": occur_krw,
            "회사적용환율(내재)": implied_rate,
            "회사계상_외환차익차손": booked_gain_loss,
        })

    return pd.DataFrame(records)


def extract_yearend_transactions(journal_df: pd.DataFrame) -> pd.DataFrame:
    """'기말평가' 구분 라인에서 거래ID별 재평가 내역 추출."""
    ye_rows = journal_df[(journal_df["구분"] == "기말평가") & (journal_df["외화금액"].notna())].copy()

    records = []
    for txn_id, group in ye_rows.groupby("전표번호(거래ID)"):
        row = group.iloc[0]
        records.append({
            "거래ID": txn_id, "결산일": row["일자"], "통화": row["통화"],
            "외화금액": row["외화금액"], "회사적용환율(기말)": row["적용환율"],
        })
    return pd.DataFrame(records)


def get_unsettled_yearend_candidates(journal_df: pd.DataFrame, schedule_df: pd.DataFrame) -> pd.DataFrame:
    """명세서(외화자산부채명세서) 기준 기말 미결제 건 중, 분개장에 기말평가 라인이
    아예 없는 거래를 찾아냄 (재평가 누락 탐지)."""
    outstanding = schedule_df[schedule_df["기말미결제외화잔액"] > 0]
    ye_done_ids = set(extract_yearend_transactions(journal_df)["거래ID"]) if not journal_df.empty else set()
    missing = outstanding[~outstanding["전표번호(거래ID)"].isin(ye_done_ids)]
    return missing


# ------------------------------------------------------------------
# 3. 1단계: 외환차손익 스크리닝
# ------------------------------------------------------------------

def screen_fx_settlements(settlements: pd.DataFrame) -> pd.DataFrame:
    """각 결제 건에 대해:
    1) 공식 매매기준율과의 '괴리율'을 계산해 5% 초과 건을 개별 플래그
    2) 공식환율 기준으로 재계산한 외환차손익과 회사 계상액의 '금액 차이(KRW)'도 함께 산출
       (개별로는 5% 이내라도, 이 금액 차이들을 합산해 중요성 검토를 하기 위함)
    회사측 환율은 실제 분개 금액에서 역산한 내재환율(원 단위, 정규화 불필요)을 사용."""
    results = []
    for _, row in settlements.iterrows():
        cur_unit = CUR_UNIT_MAP.get(row["통화"], row["통화"])
        official_raw, actual_date = get_official_rate(row["결제일"], cur_unit)
        official_rate = normalize_rate(official_raw, row["통화"])
        company_rate = row["회사적용환율(내재)"]

        deviation_pct = abs(company_rate - official_rate) / official_rate
        flagged = deviation_pct > TOLERANCE_PCT

        # 공식환율로 재계산한 결제원화금액 및 외환차손익 (부호는 booked_gain_loss와 동일 규약:
        # 채권은 (결제-발생)이 이익, 채무는 (발생-결제)가 이익)
        recalculated_settle_krw = row["외화금액"] * official_rate
        if row["구분"] == "채권":
            recalculated_gain_loss = recalculated_settle_krw - row["발생원화금액"]
        else:
            recalculated_gain_loss = row["발생원화금액"] - recalculated_settle_krw

        diff_krw = recalculated_gain_loss - row["회사계상_외환차익차손"]

        results.append({
            **row.to_dict(),
            "공식매매기준율": round(official_rate, 4),
            "공식환율기준일": actual_date.strftime("%Y-%m-%d"),
            "괴리율(%)": round(deviation_pct * 100, 2),
            "1차플래그": "이상치(정밀검증필요)" if flagged else "적정(스크리닝통과)",
            "재계산_외환차손익(공식환율기준)": round(recalculated_gain_loss),
            "회사계상액과의차이(KRW)": round(diff_krw),
        })
    return pd.DataFrame(results)


# ------------------------------------------------------------------
# 3-보조0. 거래처별 연간 요약 & 환율 추이 (분석적 절차 - 드릴다운 전 1차 검토용)
# ------------------------------------------------------------------
#
# 실제 감사에서는 분개장을 한 줄씩 보기 전에, 거래처별로 연간 외환차손익을
# 집계한 요약본과 연중 환율 추이를 먼저 보고 "이상해 보이는" 거래처/시점만
# 골라 분개장을 드릴다운한다. 아래 두 함수가 그 1차 검토 단계를 지원한다.

def build_counterparty_summary(screened: pd.DataFrame) -> pd.DataFrame:
    """거래처(+통화) 단위로 연간 결제 건수, 총 외환차손익, 회사계상액과의 차이,
    평균/최대 괴리율, 이상치 건수를 집계. 요약본에서 숫자가 튀는 거래처만
    골라 분개장을 드릴다운하면 된다."""
    if screened.empty:
        return pd.DataFrame()

    def _agg(g):
        return pd.Series({
            "결제건수": len(g),
            "총외화금액": g["외화금액"].sum(),
            "회사계상_외환차손익_합계": g["회사계상_외환차익차손"].sum(),
            "재계산_외환차손익_합계": g["재계산_외환차손익(공식환율기준)"].sum(),
            "차이_합계(KRW)": g["회사계상액과의차이(KRW)"].sum(),
            "평균괴리율(%)": round(g["괴리율(%)"].mean(), 2),
            "최대괴리율(%)": round(g["괴리율(%)"].max(), 2),
            "이상치건수": (g["1차플래그"] == "이상치(정밀검증필요)").sum(),
        })

    summary = screened.groupby(["거래처", "통화"], dropna=False).apply(_agg, include_groups=False).reset_index()
    return summary.sort_values("차이_합계(KRW)", key=abs, ascending=False)


def build_rate_trend_summary(screened: pd.DataFrame) -> pd.DataFrame:
    """월별로 회사 평균 적용환율(내재)과 공식 평균환율을 나란히 보여줘, 연중
    환율 흐름에서 이상한 시점(월)이 있는지 한눈에 볼 수 있게 한다."""
    if screened.empty:
        return pd.DataFrame()

    df = screened.copy()
    df["결제월"] = pd.to_datetime(df["결제일"]).dt.strftime("%Y-%m")

    def _agg(g):
        return pd.Series({
            "거래건수": len(g),
            "회사평균적용환율": round(g["회사적용환율(내재)"].mean(), 2),
            "공식평균환율": round(g["공식매매기준율"].mean(), 4),
            "평균괴리율(%)": round(g["괴리율(%)"].mean(), 2),
            "이상치건수": (g["1차플래그"] == "이상치(정밀검증필요)").sum(),
        })

    trend = df.groupby(["결제월", "통화"], dropna=False).apply(_agg, include_groups=False).reset_index()
    return trend.sort_values(["통화", "결제월"])


def check_aggregate_materiality(screened: pd.DataFrame, ampt: float) -> dict:
    """개별 건이 전부 5% 이내로 통과했더라도, 재계산액과 회사계상액의 차이를
    전체 합산했을 때 AMPT(허용가능 오류금액)를 넘는지 확인.
    순합계(넷)와 절대값합계(그로스) 둘 다 보여줌 - 방향이 서로 반대인 오류가
    상쇄되어 순합계는 작아 보여도 그로스 기준으로는 클 수 있기 때문."""
    net_sum = screened["회사계상액과의차이(KRW)"].sum()
    gross_sum = screened["회사계상액과의차이(KRW)"].abs().sum()
    breach = abs(net_sum) > ampt or gross_sum > ampt

    return {
        "AMPT": ampt,
        "순차이합계(KRW)": round(net_sum),
        "절대값차이합계(KRW)": round(gross_sum),
        "중요성초과여부": breach,
        "판정": "전체 재검토 필요(합계 중요성 초과)" if breach else "전체 적정(합계 중요성 이내)",
    }


# ------------------------------------------------------------------
# 3-보조. 결제일-환율기준일 불일치 탐지
# ------------------------------------------------------------------
#
# 왜 필요한가: 5% 괴리율 기준(크기 기준)은 "전월말 환율을 잘못 쓴" 것 같은 오류를
# 놓칠 수 있다 (원/달러가 한 달 새 5% 넘게 안 움직이면 개별 판정을 통과해버림 -
# TXN002 사례에서 실측). 그래서 크기가 아니라 "회사가 쓴 환율이 실제로는 다른
# 날짜의 공식 환율과 정확히 일치하는가"를 별도로 확인한다. 소액이라도 날짜
# 자체를 잘못 쓴 경우라면 이 체크에서 걸린다.

REF_DATE_MATCH_TOLERANCE = 0.0005  # 환율이 '일치'한다고 볼 허용오차(0.05%) - API 반올림 오차 흡수용


def _candidate_wrong_dates(settle_date: datetime) -> list:
    """실무에서 실제로 자주 발생하는 '기준일 착오' 패턴 후보만 골라서 반환.
    (임의로 40일을 다 훑는 대신, 흔한 실수 패턴만 확인 -> API 호출을 대폭 절감)"""
    candidates = []

    # 전월말(직전월 마지막 날)
    first_of_this_month = settle_date.replace(day=1)
    prev_month_end = first_of_this_month - timedelta(days=1)
    candidates.append(prev_month_end)

    # 전영업일들 (T-1 ~ T-3, 주말 포함해서 최대 5일 전까지)
    for i in range(1, 6):
        candidates.append(settle_date - timedelta(days=i))

    # 1주일 전, 2주일 전 (담당자가 "지난주 환율로" 착각하는 경우)
    candidates.append(settle_date - timedelta(days=7))
    candidates.append(settle_date - timedelta(days=14))

    # 중복 제거, 결제일 이후 날짜는 제외
    seen = set()
    result = []
    for d in candidates:
        key = d.strftime("%Y%m%d")
        if key not in seen and d < settle_date:
            seen.add(key)
            result.append(d)
    return result


def detect_reference_date_mismatch(settlements: pd.DataFrame) -> pd.DataFrame:
    """회사가 사용한 환율이 결제일이 아닌 다른 '흔히 착각하는' 날짜의 공식환율과
    더 정확히 일치하는지 확인. 일치하는 과거 날짜를 찾으면 '기준일 오류 의심'으로
    표시하고 그 날짜를 함께 보여준다. 5% 이내로 통과한 건도 전부 검사 대상.
    (거래당 API 호출을 최대 8회 내외로 제한 - 이전 버전은 최대 40회까지 순차 호출해서
    수출입은행 서버의 사실상 레이트리밋에 걸렸었음)"""
    results = []
    for _, row in settlements.iterrows():
        cur_unit = CUR_UNIT_MAP.get(row["통화"], row["통화"])
        company_rate = row["회사적용환율(내재)"]
        settle_date = row["결제일"]

        matched_date = None
        for candidate_date in _candidate_wrong_dates(settle_date):
            rate_table = fetch_rates_for_date(candidate_date.strftime("%Y%m%d"))
            if cur_unit not in rate_table:
                continue
            candidate_rate = normalize_rate(rate_table[cur_unit], row["통화"])
            if candidate_rate == 0:
                continue
            if abs(company_rate - candidate_rate) / candidate_rate <= REF_DATE_MATCH_TOLERANCE:
                matched_date = candidate_date
                break

        official_settle_raw, _ = get_official_rate(settle_date, cur_unit)
        official_settle_rate = normalize_rate(official_settle_raw, row["통화"])
        already_matches_settle_date = (
            official_settle_rate != 0 and
            abs(company_rate - official_settle_rate) / official_settle_rate <= REF_DATE_MATCH_TOLERANCE
        )

        if already_matches_settle_date:
            verdict = "정상(결제일 환율 사용 확인)"
            matched_date_str = None
        elif matched_date is not None:
            verdict = f"기준일 오류 의심 - {matched_date.strftime('%Y-%m-%d')} 환율을 사용한 것으로 보임"
            matched_date_str = matched_date.strftime("%Y-%m-%d")
        else:
            verdict = "불일치(흔한 패턴과 매칭 안 됨 - 은행 우대환율 등 개별 사유 가능, 증빙 확인 필요)"
            matched_date_str = None

        results.append({
            "거래ID": row["거래ID"], "결제일": settle_date.strftime("%Y-%m-%d"),
            "회사적용환율(내재)": company_rate, "기준일판정": verdict, "추정사용일자": matched_date_str,
        })

    return pd.DataFrame(results)


# ------------------------------------------------------------------
# 4. 2단계: 증빙 OCR (Claude Vision)
# ------------------------------------------------------------------

def _find_evidence_file(txn_id: str) -> str | None:
    for ext in (".png", ".jpg", ".jpeg", ".pdf"):
        candidate = os.path.join(EVIDENCE_DIR, f"{txn_id}{ext}")
        if os.path.exists(candidate):
            return candidate
    return None


def extract_rate_from_evidence(image_path: str) -> dict:
    """증빙 이미지(은행 외환거래확인서 등)에서 실제 적용환율/금액을 Claude Vision으로 추출.
    반환: {"거래일자":..., "통화":..., "적용환율":..., "외화금액":..., "원화금액":...}
    파싱 실패 시 raw_text에 원본 응답을 담아 반환하니, 로그로 확인해서 프롬프트를 조정하세요."""
    import anthropic  # 로컬 실행 시 pip install anthropic 필요

    client = anthropic.Anthropic()  # ANTHROPIC_API_KEY 환경변수 사용

    with open(image_path, "rb") as f:
        image_b64 = base64.standard_b64encode(f.read()).decode("utf-8")

    ext = os.path.splitext(image_path)[1].lower()
    media_type = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg"}.get(ext.strip("."), "image/png")

    if ext == ".pdf":
        content_block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": image_b64}}
    else:
        content_block = {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}}

    prompt = (
        "이 이미지는 은행의 외환거래확인서, SWIFT 통지서, 또는 외화예금 거래명세표입니다. "
        "여기서 실제 적용된 환율과 거래 정보를 추출하세요. "
        "다른 설명 없이 아래 형식의 JSON 객체 하나만 답하세요:\n"
        '{"거래일자": "YYYY-MM-DD", "통화": "USD", "적용환율": 1234.5, '
        '"외화금액": 10000, "원화금액": 12345000}\n'
        "값을 찾을 수 없는 항목은 null로 표기하세요."
    )

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=500,
        messages=[{"role": "user", "content": [content_block, {"type": "text", "text": prompt}]}],
    )

    raw_text = "".join(b.text for b in message.content if b.type == "text")
    cleaned = re.sub(r"```json|```", "", raw_text).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        return {"error": "JSON 파싱 실패", "raw_text": raw_text}


def verify_with_evidence(flagged_df: pd.DataFrame) -> pd.DataFrame:
    """1단계에서 플래그된 건에 대해 증빙 파일이 있으면 OCR로 검증, 없으면 '증빙요청필요' 표시."""
    rows = []
    for _, row in flagged_df.iterrows():
        record = row.to_dict()
        evidence_path = _find_evidence_file(row["거래ID"])

        if evidence_path is None:
            record["증빙상태"] = "증빙 요청 필요"
            record["증빙확인환율"] = None
            record["최종판정"] = "미확인(증빙 미확보)"
            rows.append(record)
            continue

        extracted = extract_rate_from_evidence(evidence_path)
        if "error" in extracted:
            record["증빙상태"] = "OCR 인식 실패 - 수기 확인 필요"
            record["증빙확인환율"] = None
            record["최종판정"] = "미확인(OCR 실패)"
            rows.append(record)
            continue

        confirmed_rate = extracted.get("적용환율")
        record["증빙상태"] = "증빙 확인 완료"
        record["증빙확인환율"] = confirmed_rate

        if confirmed_rate is not None:
            currency = row["통화"]
            fc = row["외화금액"]
            # 재계산: 발생시점 원가와의 차이는 별도 원장 대조가 필요하므로,
            # 여기서는 "증빙 환율 vs 회사 계상 환율" 자체의 일치 여부를 우선 확인
            company_rate = row["회사적용환율(내재)"]
            rate_match = abs(confirmed_rate - company_rate) <= max(company_rate * 0.001, 0.01)
            record["최종판정"] = "적정(증빙과 일치)" if rate_match else "부적정(증빙과 불일치 - 재계산 필요)"
        else:
            record["최종판정"] = "미확인(증빙에서 환율 추출 실패)"

        rows.append(record)

    return pd.DataFrame(rows)


# ------------------------------------------------------------------
# 5. 외화환산손익 (기말평가) - 전수 자동 검증
# ------------------------------------------------------------------

def verify_yearend_translation(ye_df: pd.DataFrame, missing_df: pd.DataFrame, year_end_date: str) -> pd.DataFrame:
    rows = []
    ye_dt = pd.to_datetime(year_end_date)

    for _, row in ye_df.iterrows():
        cur_unit = CUR_UNIT_MAP.get(row["통화"], row["통화"])
        official_raw, actual_date = get_official_rate(ye_dt, cur_unit)
        official_rate = normalize_rate(official_raw, row["통화"])
        company_rate = normalize_rate(row["회사적용환율(기말)"], row["통화"]) if row["통화"] == "JPY" else row["회사적용환율(기말)"]

        rate_match = abs(company_rate - official_rate) <= official_rate * 0.001
        rows.append({
            **row.to_dict(),
            "공식결산환율": round(official_rate, 4),
            "일치여부": "적정" if rate_match else "부적정(환율 오류)",
        })

    for _, row in missing_df.iterrows():
        rows.append({
            "거래ID": row["전표번호(거래ID)"], "결산일": year_end_date, "통화": row["통화"],
            "외화금액": row["기말미결제외화잔액"], "회사적용환율(기말)": None,
            "공식결산환율": None, "일치여부": "부적정(기말평가 누락)",
        })

    return pd.DataFrame(rows)


# ------------------------------------------------------------------
# 5-보조. 계정별원장 대사 (분개장 <-> 계정별원장 교차검증)
# ------------------------------------------------------------------
#
# 분개장과 명세서만 봐서는 못 잡는 유형의 오류가 있다: 상위 시스템에서 수기로
# 조정분개를 넣었는데 그게 상세 분개장에는 반영이 안 된 경우. 분개장 자체는
# 내부적으로 앞뒤가 맞으니 분개장만 보면 이상이 없어 보이지만, 계정별원장(총계정원장)의
# 기말잔액과 대조하면 차이가 드러난다. 그래서 외환차손익/외화환산손익 관련
# 4개 계정(외환차익/외환차손/외화환산이익/외화환산손실)에 한해 분개장에서 집계한
# 차변·대변 합계를 계정별원장 상 금액과 대사한다.

LEDGER_RECONCILE_ACCOUNTS = {
    "907": "외환차익", "957": "외환차손", "908": "외화환산이익", "958": "외화환산손실",
}
LEDGER_RECONCILE_TOLERANCE_KRW = 1000  # 반올림 오차 흡수용


def verify_ledger_reconciliation(journal_df: pd.DataFrame, ledger_df: pd.DataFrame) -> pd.DataFrame:
    """분개장에서 계정별로 집계한 차변/대변 합계가 계정별원장(총계정원장) 상
    차변합계/대변합계와 일치하는지 확인. 코드/컬럼명이 문자열-숫자 혼재 등으로
    지저분해도 최대한 강건하게 매칭하도록 계정코드를 문자열로 정규화해서 비교."""
    ledger_df = ledger_df.copy()
    ledger_df["계정코드"] = ledger_df["계정코드"].astype(str).str.strip()

    rows = []
    for code, name in LEDGER_RECONCILE_ACCOUNTS.items():
        journal_rows = journal_df[journal_df["계정코드"].astype(str).str.strip() == code]
        journal_debit = pd.to_numeric(journal_rows["차변금액"], errors="coerce").fillna(0).sum()
        journal_credit = pd.to_numeric(journal_rows["대변금액"], errors="coerce").fillna(0).sum()

        ledger_row = ledger_df[ledger_df["계정코드"] == code]
        if ledger_row.empty:
            rows.append({
                "계정코드": code, "계정과목": name,
                "분개장_차변합계": round(journal_debit), "분개장_대변합계": round(journal_credit),
                "원장_차변합계": None, "원장_대변합계": None,
                "일치여부": "부적정(계정별원장에 해당 계정 없음)",
            })
            continue

        ledger_debit = pd.to_numeric(ledger_row.iloc[0]["차변합계"], errors="coerce")
        ledger_credit = pd.to_numeric(ledger_row.iloc[0]["대변합계"], errors="coerce")

        debit_match = abs(journal_debit - ledger_debit) <= LEDGER_RECONCILE_TOLERANCE_KRW
        credit_match = abs(journal_credit - ledger_credit) <= LEDGER_RECONCILE_TOLERANCE_KRW

        rows.append({
            "계정코드": code, "계정과목": name,
            "분개장_차변합계": round(journal_debit), "분개장_대변합계": round(journal_credit),
            "원장_차변합계": round(ledger_debit), "원장_대변합계": round(ledger_credit),
            "일치여부": "적정" if (debit_match and credit_match) else "부적정(분개장-원장 불일치)",
        })

    return pd.DataFrame(rows)


# ------------------------------------------------------------------
# 6. 결과 엑셀 저장
# ------------------------------------------------------------------

FLAG_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
OK_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
HEADER_FILL_XL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BOLD_XL = Font(bold=True)

FLAG_VALUES = {"이상치(정밀검증필요)", "부적정(분개장-원장 불일치)", "부적정(환율 오류)",
               "부적정(기말평가 누락)", "부적정(계정별원장에 해당 계정 없음)",
               "전체 재검토 필요(합계 중요성 초과)", "미확인(증빙 미확보)",
               "미확인(OCR 실패)", "부적정(증빙과 불일치 - 재계산 필요)"}
OK_VALUES = {"적정(스크리닝통과)", "적정", "전체 적정(합계 중요성 이내)",
             "정상(결제일 환율 사용 확인)", "증빙 확인 완료", "적정(증빙과 일치)"}


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)
    if df is None or df.empty:
        ws.append(["(해당 없음)"])
        return
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD_XL
        cell.fill = HEADER_FILL_XL
    for _, row in df.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.tolist()])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(df.columns) + 1):
            val = ws.cell(row=r, column=c).value
            if val in FLAG_VALUES:
                ws.cell(row=r, column=c).fill = FLAG_FILL
            elif val in OK_VALUES:
                ws.cell(row=r, column=c).fill = OK_FILL
    for c in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"


def export_results_to_excel(output_path: str, *, counterparty_summary, rate_trend,
                             screened, agg, ref_check, verified, ye_result, ledger_result):
    """전체 검증 결과를 시트별로 나눠 하나의 엑셀 워크북으로 저장.
    이상치/부적정 셀은 빨간색, 적정 셀은 초록색으로 하이라이트."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws0 = wb.create_sheet("요약")
    ws0.append(["외환차손익·외화환산손익 검증 결과 요약"])
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.append([])
    ws0.append(["합계 중요성(AMPT)", agg["AMPT"]])
    ws0.append(["순차이합계(KRW)", agg["순차이합계(KRW)"]])
    ws0.append(["절대값차이합계(KRW)", agg["절대값차이합계(KRW)"]])
    ws0.append(["판정", agg["판정"]])
    if agg["판정"] in FLAG_VALUES:
        ws0.cell(row=ws0.max_row, column=2).fill = FLAG_FILL
    for r in range(3, 7):
        ws0.cell(row=r, column=1).font = BOLD_XL
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 40

    _write_sheet(wb, "0.거래처별요약", counterparty_summary)
    _write_sheet(wb, "0.월별환율추이", rate_trend)
    _write_sheet(wb, "A.외환차손익_스크리닝", screened)
    _write_sheet(wb, "A.기준일불일치탐지", ref_check)
    _write_sheet(wb, "A.증빙검증(2단계)", verified)
    _write_sheet(wb, "B.외화환산손익", ye_result)
    _write_sheet(wb, "C.계정별원장대사", ledger_result)

    wb.save(output_path)
    print(f"\n결과 엑셀 저장 완료: {output_path}")


# ------------------------------------------------------------------
# 7. 실행
# ------------------------------------------------------------------

if __name__ == "__main__":
    journal = load_journal("분개장.xlsx")
    schedule = pd.read_excel("명세서_외화자산부채명세서.xlsx")
    ledger = pd.read_excel("계정별원장.xlsx")

    # 0단계: 실무 감사 절차와 동일하게, 거래 하나하나를 보기 전에
    # 거래처별 연간 요약 + 월별 환율 추이부터 확인한다.
    settlements = extract_settlement_transactions(journal)
    screened_preview = screen_fx_settlements(settlements)  # 요약 산출을 위해 먼저 1회 계산

    print("=== 0단계. 거래처별 연간 요약 (분석적 절차 - 여기서 이상해 보이는 거래처만 드릴다운) ===")
    counterparty_summary = build_counterparty_summary(screened_preview)
    print(counterparty_summary.to_string(index=False))

    print("\n=== 0단계-보조. 월별 환율 추이 (연중 이상 시점 파악용) ===")
    rate_trend = build_rate_trend_summary(screened_preview)
    print(rate_trend.to_string(index=False))

    print("\n=== A. 외환차손익 1단계 스크리닝 (거래 단위 상세) ===")
    if settlements["데이터무결성경고"].notna().any():
        print("⚠ 데이터 무결성 경고 발견:")
        print(settlements.loc[settlements["데이터무결성경고"].notna(), ["거래ID", "데이터무결성경고"]].to_string(index=False))
    screened = screened_preview
    print(screened[["거래ID", "통화", "회사적용환율(내재)", "공식매매기준율", "괴리율(%)",
                     "1차플래그", "회사계상액과의차이(KRW)"]])

    print("\n=== A-보조. 합계 중요성 검증 (개별 통과 건 포함 전체 합산) ===")
    agg = check_aggregate_materiality(screened, AMPT)
    print(agg)

    print("\n=== A-보조2. 결제일-환율기준일 불일치 탐지 (5% 통과 건 포함 전수 검사) ===")
    ref_check = detect_reference_date_mismatch(settlements)
    print(ref_check[["거래ID", "결제일", "기준일판정", "추정사용일자"]])

    print("\n=== A. 외환차손익 2단계 증빙검증 ===")
    if agg["중요성초과여부"]:
        print("↳ 합계 중요성 초과 - 개별 통과 건까지 포함해 전체를 정밀검증 대상으로 확장합니다.")
        to_verify = screened
    else:
        to_verify = screened[screened["1차플래그"] == "이상치(정밀검증필요)"]
    verified = verify_with_evidence(to_verify)
    print(verified[["거래ID", "증빙상태", "증빙확인환율", "최종판정"]])

    print("\n=== B. 외화환산손익 전수 검증 ===")
    ye_txns = extract_yearend_transactions(journal)
    missing_ye = get_unsettled_yearend_candidates(journal, schedule)
    ye_result = verify_yearend_translation(ye_txns, missing_ye, "2025-12-31")
    print(ye_result[["거래ID", "통화", "일치여부"]])

    print("\n=== C. 계정별원장 대사 (분개장 <-> 총계정원장 교차검증) ===")
    ledger_result = verify_ledger_reconciliation(journal, ledger)
    print(ledger_result)

    export_results_to_excel(
        "검증결과.xlsx",
        counterparty_summary=counterparty_summary, rate_trend=rate_trend,
        screened=screened, agg=agg, ref_check=ref_check, verified=verified,
        ye_result=ye_result, ledger_result=ledger_result,
    )
