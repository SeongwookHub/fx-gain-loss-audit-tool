# -*- coding: utf-8 -*-
"""
'더럽고 복잡한' 실무형 분개장 샘플 생성기
- 기존 8건짜리 깔끔한 샘플과 별도로, 실무에서 실제로 발생하는 지저분함과
  더 복잡한 거래 패턴을 넣어 파이프라인의 강건성(robustness)을 테스트하기 위한 목적
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
BOLD = Font(bold=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "분개장"

# 지저분함: 실제 회사 파일처럼 맨 위에 제목 행이 있고, 그 아래 빈 행도 있음
ws.append(["삼도테크(주) 2025년 분개장 - 외화 관련 계정 발췌"])
ws.append([])

headers = ["전표번호(거래ID)", "라인번호", "일자", "구분", "계정코드", "계정과목",
           "차변금액", "대변금액", "통화", "외화금액", "적용환율", "거래처", "적요", "담당자"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=3, column=c)
    cell.font = BOLD
    cell.fill = HEADER_FILL

rows = []

# TXN101 - 정상 케이스 (실제 환율 그대로, 담당자 컬럼 등 지저분함만 있고 계산은 정확)
rows += [
    ["TXN101", 1, "2025-01-20", "발생", 108, "외상매출금(외화)", 14591000, None, "USD", 10000, 1459.10, "Everest Trading", "수출 매출 인식", "김민준"],
    ["TXN101", 2, "2025-01-20", "발생", 401, "매출", None, 14591000, "USD", 10000, 1459.10, "Everest Trading", "수출 매출 인식", "김민준"],
    ["TXN101", 3, "2025-03-05", "결제", 103, "외화예금", 14428000, None, "USD", 10000, 1442.80, "Everest Trading", "대금 회수", "김민준"],
    ["TXN101", 4, "2025-03-05", "결제", 108, "외상매출금(외화)", None, 14591000, "USD", 10000, 1459.10, "Everest Trading", "채권 상계", "김민준"],
    ["TXN101", 5, "2025-03-05", "결제", 957, "외환차손", 163000, None, "USD", None, None, "Everest Trading", "외환차손 인식", "김민준"],
]

# TXN102 - 분할결제(한 거래를 두 번에 나눠 결제) - 실무에서 흔하지만 처리 로직이
# 복잡해지는 대표 케이스. 파이프라인이 이걸 제대로 다루는지 테스트.
rows += [
    ["TXN102", 1, "2025-02-10", "발생", 108, "외상매출금(외화)", 42615000, None, "USD", 30000, 1420.50, "Delta Manufacturing", "수출 매출 인식(분할결제 예정)", "이서연"],
    ["TXN102", 2, "2025-02-10", "발생", 401, "매출", None, 42615000, "USD", 30000, 1420.50, "Delta Manufacturing", "수출 매출 인식(분할결제 예정)", "이서연"],
    ["TXN102", 3, "2025-04-01", "결제", 103, "외화예금", 21273000, None, "USD", 15000, 1418.20, "Delta Manufacturing", "1차 대금 회수(50%)", "이서연"],
    ["TXN102", 4, "2025-04-01", "결제", 108, "외상매출금(외화)", None, 21307500, "USD", 15000, 1420.50, "Delta Manufacturing", "채권 일부 상계", "이서연"],
    ["TXN102", 5, "2025-04-01", "결제", 957, "외환차손", 34500, None, "USD", None, None, "Delta Manufacturing", "외환차손 인식(1차분)", "이서연"],
    ["TXN102", 6, "2025-06-20", "결제", 103, "외화예금", 21538500, None, "USD", 15000, 1435.90, "Delta Manufacturing", "2차 대금 회수(잔금)", "이서연"],
    ["TXN102", 7, "2025-06-20", "결제", 108, "외상매출금(외화)", None, 21307500, "USD", 15000, 1420.50, "Delta Manufacturing", "채권 잔여분 상계", "이서연"],
    ["TXN102", 8, "2025-06-20", "결제", 907, "외환차익", None, 231000, "USD", None, None, "Delta Manufacturing", "외환차익 인식(2차분)", "이서연"],
]

# TXN103 - 계정과목명에 앞뒤 공백이 섞인 지저분한 케이스 (실제 ERP export에서 흔함)
# 계산은 정확하지만, 텍스트 클렌징을 안 하면 계정코드 매칭에서 문제가 생길 수 있음
rows += [
    ["TXN103", 1, "2025-05-12", "발생", 108, " 외상매출금(외화)", 13842000, None, "USD", 10000, 1384.20, "Ocean Bridge Co", "수출 매출 인식", "박지훈"],
    ["TXN103", 2, "2025-05-12", "발생", 401, "매출 ", None, 13842000, "USD", 10000, 1384.20, "Ocean Bridge Co", "수출 매출 인식", "박지훈"],
    ["TXN103", 3, "2025-07-01", "결제", 103, "외화예금", 13850000, None, "USD", 10000, 1385.00, "Ocean Bridge Co", "대금 회수", "박지훈"],
    ["TXN103", 4, "2025-07-01", "결제", 108, " 외상매출금(외화)", None, 13842000, "USD", 10000, 1384.20, "Ocean Bridge Co", "채권 상계", "박지훈"],
    ["TXN103", 5, "2025-07-01", "결제", 907, "외환차익", None, 8000, "USD", None, None, "Ocean Bridge Co", "외환차익 인식", "박지훈"],
]

# TXN104 - 외화금액이 콤마 포함 텍스트로 입력된 경우 (엑셀 셀 서식이 "텍스트"로
# 잘못 지정된 실무 흔한 실수) + 적용환율 셀이 통째로 비어있음
rows += [
    ["TXN104", 1, "2025-03-22", "발생", 108, "외상매출금(외화)", 20898000, None, "USD", "15,000", 1393.20, "Silverline Corp", "수출 매출 인식", "최유리"],
    ["TXN104", 2, "2025-03-22", "발생", 401, "매출", None, 20898000, "USD", "15,000", 1393.20, "Silverline Corp", "수출 매출 인식", "최유리"],
    ["TXN104", 3, "2025-05-30", "결제", 103, "외화예금", 21120000, None, "USD", "15,000", None, "Silverline Corp", "대금 회수 (적용환율 미기재)", "최유리"],
    ["TXN104", 4, "2025-05-30", "결제", 108, "외상매출금(외화)", None, 20898000, "USD", "15,000", 1393.20, "Silverline Corp", "채권 상계", "최유리"],
    ["TXN104", 5, "2025-05-30", "결제", 907, "외환차익", None, 222000, "USD", None, None, "Silverline Corp", "외환차익 인식", "최유리"],
]

# TXN105 - 전표번호(거래ID) 중복 재사용. TXN101과 아무 관련 없는 별개의 신규 거래인데,
# 담당자가 실수로 이전에 썼던 번호를 재사용한 경우 (전표번호 채번 실수)
rows += [
    ["TXN101", 6, "2025-09-14", "발생", 251, "외상매입금(외화)", 9840000, None, "EUR", 6000, 1640.00, "Munich Parts AG", "원재료 매입(전표번호 중복오류)", "정하늘"],
    ["TXN101", 7, "2025-09-14", "발생", 146, "원재료매입", None, 9840000, "EUR", 6000, 1640.00, "Munich Parts AG", "원재료 매입(전표번호 중복오류)", "정하늘"],
]

# TXN106 - 통화 코드가 소문자로 입력된 경우 ("usd") - 실제 시스템 간 데이터 이관 시 흔함
rows += [
    ["TXN106", 1, "2025-06-02", "발생", 108, "외상매출금(외화)", 17182500, None, "usd", 12500, 1374.60, "Nordic Freight AB", "수출 매출 인식", "한소희"],
    ["TXN106", 2, "2025-06-02", "발생", 401, "매출", None, 17182500, "usd", 12500, 1374.60, "Nordic Freight AB", "수출 매출 인식", "한소희"],
    ["TXN106", 3, "2025-08-11", "결제", 103, "외화예금", 17150000, None, "usd", 12500, 1372.00, "Nordic Freight AB", "대금 회수", "한소희"],
    ["TXN106", 4, "2025-08-11", "결제", 108, "외상매출금(외화)", None, 17182500, "usd", 12500, 1374.60, "Nordic Freight AB", "채권 상계", "한소희"],
    ["TXN106", 5, "2025-08-11", "결제", 957, "외환차손", 32500, None, "usd", None, None, "Nordic Freight AB", "외환차손 인식", "한소희"],
]

# TXN107 - 계정코드가 텍스트("108.0")로 입력되어 dtype이 불안정한 경우 + 기말 미결제
rows += [
    ["TXN107", 1, "2025-10-02", "발생", "108.0", "외상매출금(외화)", 34600000, None, "USD", 25000, 1384.00, "Baltic Export OU", "수출 매출 인식(기말 미결제)", "오지훈"],
    ["TXN107", 2, "2025-10-02", "발생", 401, "매출", None, 34600000, "USD", 25000, 1384.00, "Baltic Export OU", "수출 매출 인식(기말 미결제)", "오지훈"],
]

# TXN108 - 정상적인 기말 미결제 건 (재평가까지 정확히 계상) - false positive 확인용
rows += [
    ["TXN108", 1, "2025-11-05", "발생", 108, "외상매출금(외화)", 28754000, None, "USD", 20000, 1437.70, "Cascade Foods Inc", "수출 매출 인식", "김민준"],
    ["TXN108", 2, "2025-11-05", "발생", 401, "매출", None, 28754000, "USD", 20000, 1437.70, "Cascade Foods Inc", "수출 매출 인식", "김민준"],
    ["TXN108", 3, "2025-12-31", "기말평가", 958, "외화환산손실", 56000, None, "USD", None, None, "Cascade Foods Inc", "외화환산손실 인식", "김민준"],
    ["TXN108", 4, "2025-12-31", "기말평가", 108, "외상매출금(외화)", None, 56000, "USD", 20000, 1434.90, "Cascade Foods Inc", "외화환산손실 인식", "김민준"],
]

# TXN109 - 명백한 오류: 결제 시 회사가 완전히 엉뚱한 환율(오타로 자릿수가 밀린 값)을 사용
rows += [
    ["TXN109", 1, "2025-04-14", "발생", 108, "외상매출금(외화)", 71135000, None, "USD", 50000, 1422.70, "Redwood Systems LLC", "수출 매출 인식", "이서연"],
    ["TXN109", 2, "2025-04-14", "발생", 401, "매출", None, 71135000, "USD", 50000, 1422.70, "Redwood Systems LLC", "수출 매출 인식", "이서연"],
    ["TXN109", 3, "2025-07-09", "결제", 103, "외화예금", 7113500, None, "USD", 50000, 142.27, "Redwood Systems LLC", "대금 회수 (환율 자릿수 오기재 의심)", "이서연"],
    ["TXN109", 4, "2025-07-09", "결제", 957, "외환차손", 64021500, None, "USD", None, None, "Redwood Systems LLC", "외환차손 인식", "이서연"],
    ["TXN109", 5, "2025-07-09", "결제", 108, "외상매출금(외화)", None, 71135000, "USD", 50000, 1422.70, "Redwood Systems LLC", "채권 상계", "이서연"],
]

for r in rows:
    ws.append(r)

for col_idx in range(1, len(headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
ws.freeze_panes = "A4"

wb.save("messy_분개장.xlsx")
print(f"분개장 생성 완료: 총 {len(rows)}줄")

# ------------------------------------------------------------------
# 명세서 (외화자산부채명세서)
# ------------------------------------------------------------------
# 참고: TXN105는 분개장에는 "TXN101"로 잘못 채번되어 들어가 있지만(전표번호 중복 오류),
# 명세서는 원 거래 기록을 기준으로 정확한 별도 ID(TXN105)로 관리되고 있다고 가정.
# → 분개장과 명세서가 서로 다른 거래 식별자를 쓰는, 실무에서 실제로 발생하는
#   문서 간 불일치 상황을 재현한 것.

schedule_rows = [
    ("TXN101", "Everest Trading", "USD", "채권", "108 외상매출금(외화)",
     "2025-01-20", 10000, 1459.10, 14591000, "2025-03-05", 10000, 1442.80, 14428000, 0, "기중 결제 완료"),
    ("TXN102", "Delta Manufacturing", "USD", "채권", "108 외상매출금(외화)",
     "2025-02-10", 30000, 1420.50, 42615000, "2025-04-01", 15000, 1418.20, 21273000, 0, "분할결제 1차(50%)"),
    ("TXN102", "Delta Manufacturing", "USD", "채권", "108 외상매출금(외화)",
     "2025-02-10", 30000, 1420.50, 42615000, "2025-06-20", 15000, 1435.90, 21538500, 0, "분할결제 2차(잔금)"),
    ("TXN103", "Ocean Bridge Co", "USD", "채권", "108 외상매출금(외화)",
     "2025-05-12", 10000, 1384.20, 13842000, "2025-07-01", 10000, 1385.00, 13850000, 0, "기중 결제 완료"),
    ("TXN104", "Silverline Corp", "USD", "채권", "108 외상매출금(외화)",
     "2025-03-22", 15000, 1393.20, 20898000, "2025-05-30", 15000, 1408.00, 21120000, 0, "기중 결제 완료"),
    ("TXN105", "Munich Parts AG", "EUR", "채무", "251 외상매입금(외화)",
     "2025-09-14", 6000, 1640.00, 9840000, "", "", "", "", 6000, "기말 현재 미결제 (분개장에는 TXN101로 오채번되어 있음 - 확인 필요)"),
    ("TXN106", "Nordic Freight AB", "USD", "채권", "108 외상매출금(외화)",
     "2025-06-02", 12500, 1374.60, 17182500, "2025-08-11", 12500, 1372.00, 17150000, 0, "기중 결제 완료"),
    ("TXN107", "Baltic Export OU", "USD", "채권", "108 외상매출금(외화)",
     "2025-10-02", 25000, 1384.00, 34600000, "", "", "", "", 25000, "기말 현재 미결제"),
    ("TXN108", "Cascade Foods Inc", "USD", "채권", "108 외상매출금(외화)",
     "2025-11-05", 20000, 1437.70, 28754000, "", "", "", "", 20000, "기말 현재 미결제"),
    ("TXN109", "Redwood Systems LLC", "USD", "채권", "108 외상매출금(외화)",
     "2025-04-14", 50000, 1422.70, 71135000, "2025-07-09", 50000, 142.27, 7113500, 0, "기중 결제 완료 - 적용환율 이상치 확인 필요"),
]

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "외화자산부채명세서"
sched_headers = ["전표번호(거래ID)", "거래처", "통화", "구분(채권/채무)", "계정과목",
                  "발생일", "발생외화금액", "발생환율", "발생원화금액",
                  "결제일", "결제외화금액", "결제환율", "결제원화금액",
                  "기말미결제외화잔액", "비고"]
ws2.append(sched_headers)
for c in range(1, len(sched_headers) + 1):
    cell = ws2.cell(row=1, column=c)
    cell.font = BOLD
    cell.fill = HEADER_FILL
for r in schedule_rows:
    ws2.append(list(r))
for col_idx in range(1, len(sched_headers) + 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
ws2.freeze_panes = "A2"
wb2.save("messy_명세서_외화자산부채명세서.xlsx")
print("명세서 생성 완료")

# ------------------------------------------------------------------
# 계정별원장 - rows(분개장 전체 라인)에서 계정별 차/대변 합계를 직접 집계
# ------------------------------------------------------------------
from collections import defaultdict

debit_sum = defaultdict(float)
credit_sum = defaultdict(float)
account_names = {}

for r in rows:
    code = r[4]
    name = str(r[5]).strip()
    try:
        code_norm = str(int(float(code))).strip()  # "108.0" 같은 지저분한 값도 정규화
    except (ValueError, TypeError):
        code_norm = str(code).strip()
    account_names[code_norm] = name
    debit = r[6] or 0
    credit = r[7] or 0
    debit_sum[code_norm] += debit
    credit_sum[code_norm] += credit

opening = {"103": 38000000, "108": 9500000, "251": 5200000}

# 의도적 완전성 이슈: 외환차익(907) 계정에 분개장에 없는 수기 조정분 +75,000원을
# 총계정원장에만 반영 (상위 결산 조정 후 상세 분개장에 미반영된 상황을 재현)
LEDGER_ADJUSTMENT = {"907": 75000}

wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "계정별원장"
ledger_headers = ["계정코드", "계정과목", "기초잔액", "차변합계", "대변합계", "기말잔액"]
ws3.append(ledger_headers)
for c in range(1, len(ledger_headers) + 1):
    cell = ws3.cell(row=1, column=c)
    cell.font = BOLD
    cell.fill = HEADER_FILL

order = ["103", "108", "251", "401", "146", "907", "957", "908", "958"]
asset_like = {"103", "108", "146", "957", "958"}  # 차변 증가 성격 계정

for code in order:
    name = account_names.get(code, {
        "103": "외화예금", "108": "외상매출금(외화)", "251": "외상매입금(외화)",
        "401": "매출", "146": "원재료매입", "907": "외환차익", "957": "외환차손",
        "908": "외화환산이익", "958": "외화환산손실",
    }.get(code, code))
    op = opening.get(code, 0)
    d = debit_sum.get(code, 0)
    c_ = credit_sum.get(code, 0) + LEDGER_ADJUSTMENT.get(code, 0)
    ending = op + d - c_ if code in asset_like else op + c_ - d
    ws3.append([code, name, op, round(d), round(c_), round(ending)])

for col_idx in range(1, len(ledger_headers) + 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
ws3.freeze_panes = "A2"
wb3.save("messy_계정별원장.xlsx")
print("계정별원장 생성 완료")

